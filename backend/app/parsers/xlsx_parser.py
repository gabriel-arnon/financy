from pathlib import Path

from openpyxl import load_workbook

from app.models.enums import TransactionType
from app.schemas.common import NormalizedTransactionPreview
from app.parsers.utils import infer_transaction_type, parse_brazilian_money, parse_date


def parse(path: Path, mime_type: str | None = None) -> list[NormalizedTransactionPreview]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    previews: list[NormalizedTransactionPreview] = []

    for values in rows[1:]:
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers)}
        raw_date = row.get("data") or row.get("date") or row.get("transaction_date")
        description = row.get("descricao") or row.get("descrição") or row.get("description") or row.get("historico")
        raw_amount = row.get("valor") or row.get("amount") or row.get("value")
        if raw_date is None or description is None or raw_amount is None:
            continue
        amount = parse_brazilian_money(str(raw_amount))
        previews.append(
            NormalizedTransactionPreview(
                transaction_date=parse_date(str(raw_date)),
                description=str(description).strip(),
                original_description=str(description).strip(),
                amount=abs(amount),
                type=TransactionType(infer_transaction_type(str(description), amount)),
                raw_row={key: str(value) for key, value in row.items()},
                parser_confidence=0.88,
            )
        )
    return previews
